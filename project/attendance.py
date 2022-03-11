import os
from datetime import datetime as dt  # datetimeモジュールをdtに
import openpyxl as excel # excelファイルを操作する方法
import hashlib
from flask import make_response
from datetime import timedelta as td
from datetime import timezone as tz
import datetime
from cs50 import SQL
from flask import Flask, render_template, request, session, redirect
# ログイン機能を実装
# セッションを使えるようにする
from flask_session import Session
# パスワードチェックを使えるようにする
from werkzeug.security import check_password_hash, generate_password_hash


# SQLを使えるようにする
db = SQL("sqlite:///shift.db")

app = Flask(__name__)

# sessionを使えるようにしたい
app.secret_key = 'user'
app.permanent_session_lifetime = td(minutes=5) # -> 5分 #(days=5) -> 5日保存

# JSTを日本時間とする
JST = tz(td(hours=+9), "JST")


# 最初はログインページへ
@app.route('/')
def index():
    return render_template("login.html")


@app.route('/login', methods=["GET", "POST"])
def login():

    # セッションをなくす
    session.clear()

    if request.method == "POST":
        session.parmanent = True  # permanent sessionを作る
        user = request.form["username"]  # ユーザー情報を保存
        session["user"] = user  # sessionにuser情報を保存

        # hashで管理者のパスワードを保存
        manager_password = "Manager4649"
        manager_password = hashlib.sha256(manager_password.encode("utf-8")).hexdigest()
        # 管理者のユーザーネームとパスワードが一致したら
        if request.form.get("username") == "manager" and hashlib.sha256(request.form.get("password").encode('utf-8')).hexdigest() == manager_password:
            return render_template("management.html")

        # shifts tableから入力されたユーザーネームの情報を取り出す
        rows = db.execute("SELECT * FROM shifts WHERE username = ?", request.form.get("username"))

        # 配列がなかったら、つまりuserに関する情報が存在しなかったら
        if not rows:
            return render_template("login-again.html")

        # 入力されたパスワードをhashする→確かめられるように
        hashed_password = hashlib.sha256(request.form.get("password").encode('utf-8')).hexdigest()

        # userがいる際に、パスワードがあっているかを確認する
        # 入力されたユーザー名のデータにおいて、パスワードが入力されたものであるかをチェック
        password = db.execute("SELECT password FROM shifts WHERE username = ? AND password = ?",
        request.form.get("username"), hashed_password)

        if not password:
            return render_template("login-again.html")

        # ユーザーネームとパスワードがあっていたら、時間入力画面へ
        return render_template("input-time.html")

    else:
        return render_template("login.html")


@app.route('/register', methods=["GET", "POST"])
def register():

    if request.method == "POST":
        # パスワードと確認パスワードが違ったら、まずはやり直し
        if request.form.get("password") != request.form.get("confirmation"):
            return render_template("register-again-password.html")

        # 安全性を高めるためにパスワードをハッシュ https://qiita.com/keisuke0508/items/e11d6a441ed2691ec74a
        hashed_password = hashlib.sha256(request.form.get("password").encode('utf-8')).hexdigest()

        # 登録前に、被っている名前ないか確認
        # primary keyとしてはmailadressを使いたいから、いつか変えるかも
        existing_name = db.execute("SELECT username FROM shifts WHERE username = ?", request.form.get("username"))

        if not existing_name:
            # 名前とパスワードを登録する
            db.execute("INSERT INTO shifts (username, password) VALUES(?, ?)",
            request.form.get("username"), hashed_password)

            return redirect("/")

        # ユーザーネームの重複があったら、やり直し
        else:
            return render_template("register-again-username.html")


@app.route('/input-time', methods=["GET", "POST"])
def input_time():
    if request.method == "POST":
        return render_template("input-time.html")


@app.route('/attendance', methods=["GET", "POST"])
def attendance():
    if request.method == "POST":
        # ボタンが押されたタイミングの時間計測
        date_today = datetime.date.today()
        time_now = dt.now(JST).time()
        time_now = time_now.replace(second=0, microsecond=0)

        db.execute("UPDATE shifts SET attendance_time = ? WHERE username = ?",
        time_now, session["user"])

        # 日付を入力
        db.execute("UPDATE shifts SET date = ? WHERE username = ?",
        date_today, session["user"])

        return render_template("attendance.html", time=time_now)


@app.route('/break-start', methods=["GET", "POST"])
def break_start():
    if request.method == "POST":
        # ボタンが押されたタイミングの時間計測
        date_today = datetime.date.today()
        time_now = dt.now(JST).time()
        time_now = time_now.replace(second=0, microsecond=0)

        db.execute("UPDATE shifts SET break_start = ? WHERE username = ?",
        time_now, session["user"])

        db.execute("UPDATE shifts SET date = ? WHERE username = ?",
        date_today, session["user"])

        return render_template("break-start.html", time=time_now)


@app.route('/break-end', methods=["GET", "POST"])
def break_end():
    if request.method == "POST":
        # ボタンが押されたタイミングの時間計測
        date_today = datetime.date.today()
        time_now = dt.now(JST).time()
        time_now = time_now.replace(second=0, microsecond=0)  # 秒以下を切り捨て

        db.execute("UPDATE shifts SET break_end = ? WHERE username = ?",
        time_now, session["user"])

        db.execute("UPDATE shifts SET date = ? WHERE username = ?",
        date_today, session["user"])

        return render_template("break-end.html", time=time_now)


@app.route('/leave', methods=["GET", "POST"])
def leave():
    if request.method == "POST":
        # ボタンが押されたタイミングの時間計測
        date_today = datetime.date.today()
        time_now = dt.now(JST).time()
        time_now = time_now.replace(second=0, microsecond=0)

        db.execute("UPDATE shifts SET leave_time = ? WHERE username = ?",
        time_now, session["user"])

        db.execute("UPDATE shifts SET date = ? WHERE username = ?",
        date_today, session["user"])

        return render_template("leave.html", time=time_now)


@app.route('/calculate', methods=["GET", "POST"])
def calculate():
    if request.method == "POST":

        attendance_time = db.execute("SELECT attendance_time FROM shifts WHERE username = ? AND date = ?",
        session["user"], request.form.get("date"))[0]["attendance_time"] # [0]["attendance_time"]をつけることで、リストから一つの値を取る

        break_start = db.execute("SELECT break_start FROM shifts WHERE username = ? AND date = ?",
        session["user"], request.form.get("date"))[0]["break_start"]  # [0]だから、同じusernameが2つ以上あると使えなくなる→解決済

        break_end = db.execute("SELECT break_end FROM shifts WHERE username = ? AND date = ?",
        session["user"], request.form.get("date"))[0]["break_end"]

        leave_time = db.execute("SELECT leave_time FROM shifts WHERE username = ? AND date = ?",
        session["user"], request.form.get("date"))[0]["leave_time"]

        # 文字列を時間にする&時間、分、秒だけに
        attendance_time = dt.strptime(attendance_time, '%H:%M:%S')
        break_start = dt.strptime(break_start, '%H:%M:%S')
        break_end = dt.strptime(break_end, '%H:%M:%S')
        leave_time = dt.strptime(leave_time, '%H:%M:%S')

        # 合計勤務時間を計算
        sum_time = leave_time - attendance_time - (break_end - break_start)

        # データベースに入れる前に、文字列に戻す
        sum_time = str(sum_time)

        # 合計勤務時間をデータベースへ
        db.execute("UPDATE shifts SET sum_time = ? WHERE username = ?",
        sum_time, session["user"])

        return render_template("working_hours.html", sum_time=sum_time)

    else:
        return render_template("calculate.html")


# https://lightgauge.net/language/python/flask-excel-download/このURLを参考
@app.route('/download-excel', methods=["GET", "POST"])
def download_excel():
    if request.method == "POST":

        filename = "output.xlsx"

        # excel作成
        wb = excel.Workbook()
        ws = wb.active

        # 書き込む行、列の指定
        # row_num = 0
        # cell_num = 0

        ws['A1'].value = 'username'
        ws['B1'].value = 'date'
        ws['C1'].value = 'attendance_time'
        ws['D1'].value = 'break_start'
        ws['E1'].value = 'break_end'
        ws['F1'].value = 'leave_time'
        ws['G1'].value = 'sum_time'

        # 参照https://myafu-python.com/sqlite_excel
        # セルへのデータ書き込み
        # 名前順に並べる
        rows = db.execute("SELECT * FROM shifts WHERE date = ? ORDER BY username ASC",
        request.form.get("date"))
        for i, row in enumerate (rows):
            ws['A' + str(i+2)].value = rows[i]["username"] #名前
            ws['B' + str(i+2)].value = rows[i]["date"] #日にち
            ws['C' + str(i+2)].value = rows[i]["attendance_time"] #出勤時間
            ws['D' + str(i+2)].value = rows[i]["break_start"] #休憩開始
            ws['E' + str(i+2)].value = rows[i]["break_end"] #休憩終了
            ws['F' + str(i+2)].value = rows[i]["leave_time"] #退勤時間
            ws['G' + str(i+2)].value = rows[i]["sum_time"] #合計勤務時間

        # ワークブックをdiscに保存
        wb.save(filename)

        # ワークブックをクローズ
        wb.close()

        #MIMEタイプを設定
        XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        #Flaskのレスポンスを作成する
        response = make_response()

        #ダウンロードするEXCELファイルを開く
        wb = open( filename , "rb" )

        #ダウンロードするEXCELファイルをレスポンスに設定する
        response.data = wb.read()

        #ダウンロードするEXCELファイルをクローズ
        wb.close()

        #ダウンロードするEXCELファイル名を設定
        response.headers[ "Content-Disposition" ] = "attachment; filename=" + filename

        #MIMEタイプをレスポンスに設定
        response.mimetype = XLSX_MIMETYPE

        #MIMEタイプをレスポンスに設定
        os.remove

        #Flaskにレスポンスを返す
        return response


if __name__ == "__main__":
    app.run()